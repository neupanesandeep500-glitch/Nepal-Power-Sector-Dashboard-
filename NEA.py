"""
NEA.py — Nepal Electricity Authority Operational Data module
==============================================================
Everything related to the "🏭 NEA Operational Data" and
"🔬 NEA Forecast Lab" tabs lives in this one file, kept deliberately
separate from app.py so the two systems (power-plant licensing
dashboard vs. NEA operational analytics) never entangle each other.

WHAT THIS FILE DOES
--------------------
1. LIVE SYNC — pulls the source workbook straight from the Google Sheet
   the user shares (same `download_google_sheet_xlsx` helper the main
   app already uses for the power-plant data, so it's a proven path),
   on a background timer, with a cached-copy + bundled-snapshot fallback
   so the dashboard is never blank even if the network/sheet is
   unavailable.
2. DYNAMIC PARSING — nothing about fiscal years is hardcoded. Every
   sheet is parsed by *discovering* its FY columns/rows at parse time
   (scanning headers for non-empty cells), so adding a new fiscal-year
   column (or a new month block) to the Google Sheet and waiting for
   the next sync is enough to make it show up everywhere: the charts,
   the KPI marquee, and the forecast models. Nothing needs to be
   redeployed for a routine data update.
3. FORECASTING — Linear Regression, Holt Exponential Smoothing, Moving
   Average, auto-order ARIMA, SARIMA (seasonal, monthly series only),
   and a Linear+ARIMA Hybrid ensemble, all computed server-side with
   statsmodels against whatever the live dataset currently holds.
4. UNIT-AWARE ECONOMICS — Rs./unit (Rs./kWh) rates for import, export,
   and average revenue, computed with the correct unit handling (see
   `unit_economics()` docstring below) instead of the static/placeholder
   marquee figure the first draft of the dashboard shipped with.

HOW TO WIRE THIS INTO app.py
-----------------------------
    import NEA

    NEA.bootstrap()                 # call once at startup (non-blocking)
    NEA.start_background_refresh()  # call once at startup

    @server.route("/nea-operational-dashboard")
    def serve_nea_dashboard():
        return NEA.render_dashboard_html()

    # inside render_tab(), for tab == "nea_operational":
    #     html.Iframe(src="/nea-operational-dashboard", ...)

    # a second tab, "nea_forecast_lab", can call NEA.run_forecast(...)
    # and NEA.unit_economics() directly and plot the result with
    # dcc.Graph / plotly — see the __main__ smoke test at the bottom
    # of this file for the exact call shapes.
"""

from __future__ import annotations

import io
import itertools
import json
import os
import re
import threading
import traceback
import warnings
from dataclasses import dataclass, field
from datetime import datetime

import numpy as np
import openpyxl
import pandas as pd

warnings.filterwarnings("ignore")

from statsmodels.tsa.arima.model import ARIMA
from statsmodels.tsa.holtwinters import Holt
from statsmodels.tsa.statespace.sarimax import SARIMAX

_HERE = os.path.dirname(os.path.abspath(__file__))

# ── Configuration ────────────────────────────────────────────────────────

_PLACEHOLDER_SHEET_URL = (
    "https://docs.google.com/spreadsheets/d/1PzTJmKWfBe2_mXFgXZlsOhMcxP85q8C7VnczGxaMM2U/edit?usp=sharing"
)
# Kept for backward compatibility with any code importing NEA.DEFAULT_SHEET_URL
# directly. This is resolved ONCE at import time from the env var only — use
# _resolve_sheet_url() everywhere else, since that also picks up a sheet URL
# saved from the admin panel (see set_sheet_url() below) without needing a
# process restart.
DEFAULT_SHEET_URL = os.environ.get("NEA_SHEET_URL", _PLACEHOLDER_SHEET_URL)

CACHE_WORKBOOK_PATH = os.path.join(_HERE, "nea_workbook_cache.xlsx")
TEMPLATE_PATH = os.path.join(_HERE, "nea_assets", "nea_operational_dashboard_template.html")
FORECAST_TEMPLATE_PATH = os.path.join(_HERE, "nea_assets", "nea_forecast_lab_template.html")
AUTO_REFRESH_HOURS = float(os.environ.get("NEA_AUTO_REFRESH_HOURS", "6"))

# Sheet URL entered via the /admin panel is persisted here so it survives a
# process restart (as long as DATA_DIR/the app directory itself is on
# persistent storage) without needing NEA_SHEET_URL to be set as an actual
# platform environment variable. See _resolve_sheet_url()/set_sheet_url().
_NEA_CONFIG_PATH = os.path.join(os.environ.get("DATA_DIR", _HERE), "nea_config.json")


def _load_persisted_sheet_url():
    try:
        with open(_NEA_CONFIG_PATH, "r", encoding="utf-8") as f:
            return (json.load(f).get("sheet_url") or "").strip() or None
    except Exception:
        return None


def _save_persisted_sheet_url(url):
    try:
        os.makedirs(os.path.dirname(_NEA_CONFIG_PATH) or ".", exist_ok=True)
        with open(_NEA_CONFIG_PATH, "w", encoding="utf-8") as f:
            json.dump({"sheet_url": url}, f)
    except Exception:
        traceback.print_exc()


def _resolve_sheet_url():
    """Precedence: URL saved via the admin panel > NEA_SHEET_URL env var >
    the placeholder sheet baked into this file (which is NOT the user's
    data and will fail unless they happen to have access to it)."""
    return (_load_persisted_sheet_url()
            or os.environ.get("NEA_SHEET_URL")
            or _PLACEHOLDER_SHEET_URL)


def current_sheet_url() -> str:
    """The sheet URL that will actually be used on the next sync — for
    display in the admin panel."""
    return _resolve_sheet_url()

_MONTHS_BS = ["Shrawan", "Bhadra", "Ashwin", "Kartik", "Mangsir", "Poush",
              "Magh", "Falgun", "Chaitra", "Baishakh", "Jestha", "Ashadh"]

# Expected sheet names in the source workbook. Matched case/whitespace-
# insensitively (`_norm`) so a stray trailing space in the Google Sheet
# tab name (the source workbook has several) never breaks the sync.
_SHEET_ALIASES = {
    "system_loss": ["system loss"],
    "energy_balance": ["energy balance in gwh"],
    "capacity_balance": ["capacity balance in mw"],
    "energy_export": ["energy export in gwh"],
    "energy_import": ["energy import in gwh from india"],
    "annual_energy_peak": ["annual energy and peak load"],
    "consumers_growth": ["consumers growth"],
    "sales_revenue": ["sales revenue"],
    "transmission_line": ["transmission line length"],
    "substation_capacity": ["substation capacity"],
    "financial_data": ["financial data"],
}


def _norm(s):
    return re.sub(r"\s+", " ", str(s or "")).strip().lower()


def _slug(s):
    """Lowercase, alphanumeric-only fingerprint of a header/label string
    — robust to spacing, punctuation, unit-suffix, and ampersand/dash
    variations ('NEA ROR & PROR' and 'NEA ROR&PROR' both slug to
    'nearorpror'), so column matching survives the kind of small
    formatting drift that's routine in a hand-maintained Google Sheet."""
    return re.sub(r"[^a-z0-9]", "", str(s or "").lower())


def _find_key(categories: dict, must_include, must_exclude=()):
    """Return the first key in `categories` whose slug contains every
    keyword in `must_include` and none in `must_exclude`."""
    for k in categories:
        s = _slug(k)
        if all(m in s for m in must_include) and not any(x in s for x in must_exclude):
            return k
    return None


# ── Dynamic (non-hardcoded) sheet parsing ───────────────────────────────

def _sheet_by_alias(wb, key):
    wanted = _SHEET_ALIASES[key]
    for name in wb.sheetnames:
        if _norm(name) in wanted:
            return wb[name]
    return None


def _rows(ws):
    return list(ws.iter_rows(values_only=True))


def _clean_num(v):
    if v in (None, "-", "–", ""):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _extract_row_oriented_fy(ws, fy_as_int=True):
    """For sheets shaped [Fiscal Year | cat1 | cat2 | ...] with one row
    per fiscal year (System Loss, Financial Data). Discovers however
    many FY rows and category columns actually exist — no fixed count."""
    rows = _rows(ws)
    if not rows:
        return {"fy": [], "categories": {}}
    header = [str(h).strip() if h else "" for h in rows[0]]
    cat_names = [h for h in header[1:] if h]
    out_fy, out_cats = [], {c: [] for c in cat_names}
    for r in rows[1:]:
        if r[0] is None or str(r[0]).strip() == "":
            continue
        fy_val = r[0]
        if fy_as_int:
            try:
                fy_val = int(fy_val)
            except (TypeError, ValueError):
                continue
        out_fy.append(fy_val)
        for i, cat in enumerate(cat_names, start=1):
            out_cats[cat].append(_clean_num(r[i]) if i < len(r) else None)
    return {"fy": out_fy, "categories": out_cats}


def _extract_column_oriented_fy(ws):
    """For sheets shaped [Particulars | fy1 | fy2 | ... ] with one row
    per category and one column per fiscal year (Annual Energy & Peak,
    Consumers Growth, Sales Revenue). Discovers the FY list from the
    header row directly, however many columns are present."""
    rows = _rows(ws)
    if not rows:
        return {"fy": [], "categories": {}}
    header = rows[0]
    fy_list = []
    for h in header[1:]:
        if h is None or str(h).strip() == "":
            continue
        try:
            fy_list.append(int(h))
        except (TypeError, ValueError):
            fy_list.append(str(h).strip())
    out_cats = {}
    for r in rows[1:]:
        cat = r[0]
        if not cat:
            continue
        cat = str(cat).strip()
        vals = [_clean_num(v) for v in r[1:1 + len(fy_list)]]
        out_cats[cat] = vals
    return {"fy": fy_list, "categories": out_cats}


def _extract_simple_fy_columns(ws, col_names):
    """For sheets shaped [FY | val1 | val2 | ...] with a fixed small set
    of numeric columns (Transmission Line Length, Substation Capacity).
    FY here is a BS-style string ('2081/082'), kept as text."""
    rows = _rows(ws)
    out = {"fy": []}
    for c in col_names:
        out[c] = []
    for r in rows[1:]:
        if r[0] is None or str(r[0]).strip() == "":
            continue
        out["fy"].append(str(r[0]).strip())
        for i, c in enumerate(col_names, start=1):
            out[c].append(_clean_num(r[i]) if i < len(r) else None)
    return out


def _extract_monthly_by_fy(ws):
    """For Energy Export / Energy Import: [FY | 12 month columns].
    Discovers however many FY rows (and however many month columns,
    though it's normally 12) actually exist."""
    rows = _rows(ws)
    out = {}
    for r in rows[1:]:
        if r[0] is None or str(r[0]).strip() == "":
            continue
        fy = str(r[0]).strip()
        vals = [_clean_num(v) for v in r[1:] if v is not None]
        out[fy] = vals
    return out


def _extract_monthly_grouped_by_fy(ws):
    """For Energy Balance / Capacity Balance: a wide grid where the
    header row groups 12 month-columns under each FY label (row 2 has
    the actual month names, row 1 has the FY spanning every 12th cell).
    Discovers the FY groups by detecting where the FY header changes,
    so it copes with any number of FY blocks, not just four."""
    rows = _rows(ws)
    if len(rows) < 3:
        return {"fy_order": [], "categories": {}}
    fy_header = rows[0]
    month_header = rows[1]
    groups = []  # (fy_label, start_col, end_col) inclusive, 0-indexed
    cur_fy, start = None, None
    for c in range(1, len(fy_header)):
        if fy_header[c] not in (None, ""):
            if cur_fy is not None:
                groups.append((cur_fy, start, c - 1))
            cur_fy = str(fy_header[c]).strip()
            start = c
    if cur_fy is not None:
        groups.append((cur_fy, start, len(fy_header) - 1))

    categories = {}
    for r in rows[2:]:
        cat = r[0]
        if not cat:
            continue
        cat = str(cat).strip()
        by_fy = {}
        for fy, s, e in groups:
            vals = [_clean_num(v) for v in r[s:e + 1]]
            vals = [v for v in vals if v is not None]
            by_fy[fy] = vals
        categories[cat] = by_fy

    # chronological order = reverse of sheet order (sheet lists newest first)
    fy_order_chrono = list(reversed([g[0] for g in groups]))
    return {"fy_order": fy_order_chrono, "categories": categories,
            "months": [m for m in month_header[groups[0][1]:groups[0][2] + 1]] if groups else _MONTHS_BS}


def parse_workbook(path) -> dict:
    """Parse the whole NEA workbook into raw (dynamic, un-hardcoded)
    structures. Every downstream shape (dashboard JSON, forecast series,
    unit economics) is derived from this — nothing here assumes a fixed
    number of fiscal years."""
    wb = openpyxl.load_workbook(path, data_only=True)
    parsed = {}

    ws = _sheet_by_alias(wb, "system_loss")
    if ws is not None:
        d = _extract_row_oriented_fy(ws, fy_as_int=True)
        parsed["system_loss"] = d

    ws = _sheet_by_alias(wb, "financial_data")
    if ws is not None:
        d = _extract_row_oriented_fy(ws, fy_as_int=True)
        parsed["financial_data"] = d

    ws = _sheet_by_alias(wb, "annual_energy_peak")
    if ws is not None:
        parsed["annual_energy_peak"] = _extract_column_oriented_fy(ws)

    ws = _sheet_by_alias(wb, "consumers_growth")
    if ws is not None:
        parsed["consumers_growth"] = _extract_column_oriented_fy(ws)

    ws = _sheet_by_alias(wb, "sales_revenue")
    if ws is not None:
        parsed["sales_revenue"] = _extract_column_oriented_fy(ws)

    ws = _sheet_by_alias(wb, "transmission_line")
    if ws is not None:
        parsed["transmission_line"] = _extract_simple_fy_columns(
            ws, ["kv66", "kv132", "kv220", "kv400", "total", "increment"])

    ws = _sheet_by_alias(wb, "substation_capacity")
    if ws is not None:
        parsed["substation_capacity"] = _extract_simple_fy_columns(
            ws, ["capacity", "increment"])

    ws = _sheet_by_alias(wb, "energy_export")
    if ws is not None:
        parsed["energy_export"] = _extract_monthly_by_fy(ws)

    ws = _sheet_by_alias(wb, "energy_import")
    if ws is not None:
        parsed["energy_import"] = _extract_monthly_by_fy(ws)

    ws = _sheet_by_alias(wb, "energy_balance")
    if ws is not None:
        parsed["energy_balance"] = _extract_monthly_grouped_by_fy(ws)

    ws = _sheet_by_alias(wb, "capacity_balance")
    if ws is not None:
        parsed["capacity_balance"] = _extract_monthly_grouped_by_fy(ws)

    missing = [k for k in _SHEET_ALIASES if k not in parsed]
    if missing:
        raise ValueError(f"NEA workbook is missing expected sheet(s): {missing}. "
                          f"Found tabs: {wb.sheetnames}")
    return parsed


# ── Shape the parsed data into the dashboard's DATA JSON ────────────────

# Keyword groups for fuzzy-matching Energy/Capacity Balance row labels.
# Order matters: more specific entries (e.g. "monthly system energy
# demand") must be matched before their generic substrings ("import"
# alone would otherwise never conflict here, but this keeps the pattern
# consistent with the rest of the file).
_EB_KEYWORDS = {
    "ipp": (["ipp"], []),
    "nea_sub": (["nea", "subsidiary"], []),
    "nea_ror": (["ror"], []),
    "import": (["import"], []),
    "nea_storage": (["nea", "storage"], []),
    "nea_solar": (["nea", "solar"], []),
    "thermal": (["thermal"], []),
    "interruption": (["interruption"], []),
    "system_demand": (["monthly", "system", "energy", "demand"], []),
    "export": (["export"], []),
    "national_demand": (["monthly", "national", "energy", "demand"], []),
}
_CB_KEYWORDS = {
    "ipp": (["ipp"], []),
    "nea_sub": (["nea", "subsidiary"], []),
    "nea_ror": (["ror"], []),
    "import": (["import"], []),
    "nea_storage": (["nea", "storage"], []),
    "interruption": (["interruption"], []),
    "national_peak": (["monthly", "national", "peak"], []),
    "export": (["export"], []),
    "system_peak": (["monthly", "system", "peak"], []),
}
# Keyword groups (all must appear in the slugged header/row-label) used
# to fuzzy-match the Annual Energy & Peak Load categories. Written as
# keyword sets rather than exact strings so headers like "Power Purchase
# – NEA Sub" or "Total Availability (MU)" match without needing the
# source workbook's exact punctuation/wording to stay frozen forever.
_AE_KEYWORDS = {
    "nea_own": (["nea", "own"], []),
    "nea_sub": (["nea", "sub"], []),
    "ipp": (["ipp"], []),
    "india": (["india"], []),
    "total": (["total", "availab"], []),
    "national_peak": (["national", "peak"], []),
    "system_peak": (["system", "peak"], []),
}


def _series_by_alias(cats: dict, length: int):
    """Fuzzy-match each Annual Energy & Peak Load output key against the
    sheet's actual row labels, falling back to zeros so a missing/renamed
    row degrades gracefully instead of KeyError-ing the whole dashboard."""
    out = {}
    for key, (must_include, must_exclude) in _AE_KEYWORDS.items():
        found = _find_key(cats, must_include, must_exclude)
        out[key] = cats.get(found) if found else [0] * length
    return out


def build_dashboard_data(parsed: dict) -> dict:
    """Turn the raw parsed workbook into the exact JSON shape the
    dashboard's front-end JS expects (`const DATA = {...}`), computing
    every KPI/derived figure from whatever the *latest* two data points
    currently are — so a new fiscal-year row added upstream flows
    through to the marquee and the charts automatically, no code change
    needed."""
    sl = parsed["system_loss"]
    fin = parsed["financial_data"]
    ae = parsed["annual_energy_peak"]
    cg = parsed["consumers_growth"]
    sr = parsed["sales_revenue"]
    tx = parsed["transmission_line"]
    ss = parsed["substation_capacity"]
    exp = parsed["energy_export"]
    imp = parsed["energy_import"]
    eb = parsed["energy_balance"]
    cb = parsed["capacity_balance"]

    years_str = [str(y) for y in sl["fy"]]
    n = len(years_str)

    sl_cats = sl["categories"]
    trans_key = _find_key(sl_cats, ["transmission"])
    dist_key = _find_key(sl_cats, ["distribution"])
    sys_key = _find_key(sl_cats, ["system"])
    system_loss = {
        "years": years_str,
        "transmission": sl_cats.get(trans_key, [None] * n),
        "distribution": sl_cats.get(dist_key, [None] * n),
        "system": sl_cats.get(sys_key, [None] * n),
    }

    ae_years = [str(y) for y in ae["fy"]]
    ae_series = _series_by_alias(ae["categories"], len(ae_years))
    annual_energy = {"years": ae_years, **ae_series}

    cg_years = [str(y) for y in cg["fy"]]
    total_key = next((k for k in cg["categories"] if _norm(k) == "total consumers"), None)
    growth_key = next((k for k in cg["categories"] if "growth" in _norm(k)), None)
    cg_categories = {k: v for k, v in cg["categories"].items()
                     if k not in (total_key, growth_key)}
    consumers = {
        "years": cg_years,
        "total": cg["categories"].get(total_key, []) if total_key else [],
        "growth_pct": cg["categories"].get(growth_key, []) if growth_key else [],
        "categories": cg_categories,
    }

    sr_years = [str(y) for y in sr["fy"]]
    sr_total_key = next((k for k in sr["categories"] if _norm(k) in
                          ("total gross revenue", "total")), None)
    sr_growth_key = next((k for k in sr["categories"] if "growth" in _norm(k)), None)
    sr_categories = {k: v for k, v in sr["categories"].items()
                     if k not in (sr_total_key, sr_growth_key)}
    sales = {
        "years": sr_years,
        "total": sr["categories"].get(sr_total_key, []) if sr_total_key else [],
        "growth_pct": sr["categories"].get(sr_growth_key, []) if sr_growth_key else [],
        "categories": sr_categories,
    }

    fin_years = [str(y) for y in fin["fy"]]
    fin_cats = fin["categories"]
    revenue_key = _find_key(fin_cats, ["overall", "revenue"])
    profit_key = _find_key(fin_cats, ["profit"])
    import_mu_key = _find_key(fin_cats, ["import", "mu"])
    import_rs_key = (_find_key(fin_cats, ["import", "rs"]) or _find_key(fin_cats, ["import", "million"]))
    export_mu_key = _find_key(fin_cats, ["export", "mu"])
    export_rs_key = (_find_key(fin_cats, ["export", "rs"]) or _find_key(fin_cats, ["export", "million"]))
    financial = {
        "years": fin_years,
        "revenue": fin_cats.get(revenue_key, []),
        "profit_loss": fin_cats.get(profit_key, []),
        "import_mu": fin_cats.get(import_mu_key, []),
        "import_rs": fin_cats.get(import_rs_key, []),
        "export_mu": fin_cats.get(export_mu_key, []),
        "export_rs": fin_cats.get(export_rs_key, []),
    }

    transmission = {
        "years": tx["fy"], "kv66": tx["kv66"], "kv132": tx["kv132"],
        "kv220": tx["kv220"], "kv400": tx["kv400"],
        "total": tx["total"], "increment": tx["increment"],
    }
    substation = {"years": ss["fy"], "capacity": ss["capacity"], "increment": ss["increment"]}

    export_months = _MONTHS_BS
    export_data = imp_data = {}
    export_data = {fy: vals for fy, vals in exp.items()}
    import_data = {fy: vals for fy, vals in imp.items()}

    def _shape_monthly_grouped(grouped, keyword_map):
        # resolve each output key to the sheet's actual row label once
        resolved = {}
        for key, (must_include, must_exclude) in keyword_map.items():
            found = _find_key(grouped["categories"], must_include, must_exclude)
            resolved[key] = found
        out = {}
        for fy in grouped["fy_order"]:
            entry = {"months": grouped.get("months", _MONTHS_BS)}
            for key, label in resolved.items():
                series = grouped["categories"].get(label, {}) if label else {}
                entry[key] = series.get(fy, [])
            out[fy] = entry
        return out

    energy_balance_monthly = _shape_monthly_grouped(eb, _EB_KEYWORDS)
    capacity_balance_monthly = _shape_monthly_grouped(cb, _CB_KEYWORDS)

    # ── KPIs, computed from whatever the latest 2 data points are ──────
    def _pct_change(series):
        vals = [v for v in series if v is not None]
        if len(vals) < 2 or not vals[-2]:
            return None
        return round((vals[-1] - vals[-2]) / abs(vals[-2]) * 100, 2)

    def _last(series, default=0):
        vals = [v for v in series if v is not None]
        return vals[-1] if vals else default

    consumers_total = consumers["total"]
    kpi = {
        "consumer_increase": (int(consumers_total[-1] - consumers_total[-2])
                               if len(consumers_total) >= 2 else 0),
        "consumer_increase_pct": _pct_change(consumers_total) or 0,
        "revenue_growth": _pct_change(financial["revenue"]) or 0,
        "latest_profit": _last(financial["profit_loss"]),
        "profit_change": (financial["profit_loss"][-1] - financial["profit_loss"][-2]
                           if len(financial["profit_loss"]) >= 2 else 0),
        "profit_change_pct": _pct_change(financial["profit_loss"]) or 0,
        "latest_system_loss": _last(system_loss["system"]),
        "loss_reduction": (round((system_loss["system"][-2] - system_loss["system"][-1]), 2)
                            if len(system_loss["system"]) >= 2 else 0),
        "latest_total_avail": _last(annual_energy.get("total", [])),
        "avail_growth": _pct_change(annual_energy.get("total", [])) or 0,
        "latest_peak": _last(annual_energy.get("national_peak", [])),
        "peak_growth": _pct_change(annual_energy.get("national_peak", [])) or 0,
        "latest_revenue": _last(financial["revenue"]),
        "total_consumers": _last(consumers_total),
    }

    return {
        "systemLoss": system_loss, "annualEnergy": annual_energy,
        "consumers": consumers, "sales": sales, "financial": financial,
        "transmission": transmission, "substation": substation,
        "exportData": export_data, "importData": import_data,
        "energyBalanceMonthly": energy_balance_monthly,
        "capacityBalanceMonthly": capacity_balance_monthly,
        "kpi": kpi, "exportMonths": export_months,
    }


# ── Bundled fallback snapshot (used only if a live parse has never
#    succeeded — e.g. first boot with no network) ───────────────────────

def _load_bundled_fallback() -> dict:
    """If nea_assets/nea_workbook_fallback.xlsx exists (an .xlsx snapshot
    committed to the repo), parse and return it so the dashboard still
    renders even on first boot with no network/no successful sync yet.
    NOTE: as of this commit that file is NOT present in nea_assets/, so
    this currently returns {} whenever live sync fails before any sync
    has ever succeeded — that's the actual cause of a fully blank NEA
    dashboard. Fix: either commit a real nea_workbook_fallback.xlsx here,
    or get one successful sync/upload in via the admin panel (which then
    caches to nea_workbook_cache.xlsx and is reused on subsequent
    failures — see refresh())."""
    fallback_path = os.path.join(_HERE, "nea_assets", "nea_workbook_fallback.xlsx")
    if os.path.exists(fallback_path):
        return build_dashboard_data(parse_workbook(fallback_path))
    return {}


# ── Live sync / cache ────────────────────────────────────────────────────

_CACHE = {"data": None, "parsed": None, "last_sync": None, "source": None, "error": None}
_lock = threading.Lock()


def _download_google_sheet_xlsx(url_or_id, out_path):
    """Download the source workbook and validate the response is actually
    an .xlsx (not an HTML error/sign-in page), trying the Sheets export
    endpoint first and falling back to the direct Drive download endpoint."""
    import urllib.request
    import urllib.error

    sheet_id = url_or_id
    if "/" in url_or_id:
        for pat in (r"/spreadsheets/d/([a-zA-Z0-9-_]+)",
                    r"/file/d/([a-zA-Z0-9-_]+)",
                    r"id=([a-zA-Z0-9-_]+)"):
            m = re.search(pat, url_or_id)
            if m:
                sheet_id = m.group(1)
                break

    candidate_urls = [
        f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx",
        f"https://drive.google.com/uc?export=download&id={sheet_id}",
    ]

    last_err = None
    for export_url in candidate_urls:
        try:
            req = urllib.request.Request(export_url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=120) as resp:
                content = resp.read()
                ctype = resp.headers.get("Content-Type", "")
        except urllib.error.HTTPError as e:
            last_err = f"HTTP {e.code} from {export_url}"
            continue
        except Exception as e:
            last_err = f"{type(e).__name__}: {e} ({export_url})"
            continue

        # A real .xlsx is a zip archive — first two bytes are "PK". If Google
        # couldn't serve the actual file (wrong permissions, sign-in wall,
        # not a spreadsheet) it silently returns an HTML page instead.
        if content[:2] != b"PK":
            snippet = content[:200].decode("utf-8", errors="replace")
            last_err = (f"{export_url} did not return a valid .xlsx "
                        f"(Content-Type={ctype!r}; response started with: {snippet!r})")
            continue

        with open(out_path, "wb") as f:
            f.write(content)
        return out_path

    raise RuntimeError(
        f"Could not download the NEA workbook (id={sheet_id}) from any known URL. "
        f"Last error: {last_err}. Confirm the sheet is shared as 'Anyone with the "
        f"link can view' and that NEA_SHEET_URL points at it."
    )

def refresh(sheet_url: str = None) -> bool:
    """Pull the latest workbook, reparse, rebuild the dashboard JSON, and
    swap it into the live cache atomically. Returns True on success.
    Never raises — failures are recorded in _CACHE['error'] and the
    previous good data (or the bundled fallback) keeps serving."""
    sheet_url = sheet_url or _resolve_sheet_url()
    with _lock:
        try:
            _download_google_sheet_xlsx(sheet_url, CACHE_WORKBOOK_PATH)
            parsed = parse_workbook(CACHE_WORKBOOK_PATH)
            data = build_dashboard_data(parsed)
            _CACHE.update(data=data, parsed=parsed,
                          last_sync=datetime.now().strftime("%Y-%m-%d %H:%M"),
                          source="Google Sheet (live sync)", error=None)
            return True
        except Exception as exc:
            traceback.print_exc()
            _CACHE["error"] = str(exc)
            # fall back to last-known cached workbook file on disk, if any
            if _CACHE["data"] is None and os.path.exists(CACHE_WORKBOOK_PATH):
                try:
                    parsed = parse_workbook(CACHE_WORKBOOK_PATH)
                    _CACHE.update(data=build_dashboard_data(parsed), parsed=parsed,
                                  source="Cached workbook (last good sync)")
                except Exception:
                    traceback.print_exc()
            # last resort: bundled fallback snapshot (only claim this as the
            # source if it actually returned something — an empty {} means
            # nea_assets/nea_workbook_fallback.xlsx isn't present, and we'd
            # rather leave _CACHE['data'] as None so the real sync error
            # above is what gets shown, not a misleading "fallback" label)
            if _CACHE["data"] is None:
                try:
                    fallback_data = _load_bundled_fallback()
                    if fallback_data:
                        _CACHE["data"] = fallback_data
                        _CACHE["source"] = "Bundled fallback snapshot"
                except Exception:
                    traceback.print_exc()
            return False


def set_sheet_url(url: str) -> bool:
    """Called from the admin panel's 'NEA Data Sync' card. Persists the URL
    (so it's still used after a restart) and immediately attempts a sync
    with it. Raises ValueError if url is blank; sync failures are NOT
    raised — check sync_status()['error'] afterwards, same as refresh()."""
    url = (url or "").strip()
    if not url:
        raise ValueError("Please provide a Google Sheet URL or ID for the NEA data")
    _save_persisted_sheet_url(url)
    return refresh(url)


def has_persisted_sheet_url() -> bool:
    """True if a sheet URL was saved via the admin panel. While this is
    true, _resolve_sheet_url() uses it and ignores NEA_SHEET_URL — this is
    why changing the Render env var alone can look like it "does nothing"
    once the admin panel has been used once."""
    return _load_persisted_sheet_url() is not None


def clear_persisted_sheet_url() -> bool:
    """Remove the admin-saved override so _resolve_sheet_url() falls back
    to the NEA_SHEET_URL environment variable (or the placeholder) again,
    and re-syncs immediately against whichever of those is now in effect."""
    try:
        if os.path.exists(_NEA_CONFIG_PATH):
            os.remove(_NEA_CONFIG_PATH)
    except Exception:
        traceback.print_exc()
    return refresh(_resolve_sheet_url())


def bootstrap():
    """Call once at app startup. Non-blocking-ish: does one synchronous
    attempt (so the very first page load already has real data if the
    network is up), then hands off to the background timer."""
    persisted_url = _load_persisted_sheet_url()
    env_url = os.environ.get("NEA_SHEET_URL")
    if persisted_url:
        print(f"[NEA DEBUG] Using sheet URL saved via /admin ({persisted_url[:60]}...) — attempting sync.")
    elif env_url:
        print(f"[NEA DEBUG] NEA_SHEET_URL is set ({env_url[:60]}...) — attempting sync.")
    else:
        print("[NEA DEBUG] No NEA sheet configured — NEA_SHEET_URL is not set and nothing has "
              "been saved via the admin panel. Falling back to the placeholder sheet baked "
              "into NEA.py, which will almost certainly fail unless that placeholder sheet "
              "happens to be shared with you. Set NEA_SHEET_URL, or paste your sheet's URL "
              "into the 'NEA Data Sync' card on /admin.")
    if not os.path.exists(os.path.join(_HERE, "nea_assets", "nea_workbook_fallback.xlsx")):
        print("[NEA DEBUG] Note: nea_assets/nea_workbook_fallback.xlsx is not present in this "
              "deployment, so if the live sync fails and no workbook has been cached/uploaded "
              "yet, the NEA dashboard will have no data to show at all until a sync succeeds.")

    def _bootstrap_and_report():
        ok = refresh()
        status = sync_status()
        if ok:
            print(f"[NEA DEBUG] Initial sync succeeded: {status['source']} at {status['last_sync']}")
        else:
            print(f"[NEA DEBUG] Initial sync FAILED: {status['error']}")

    threading.Thread(target=_bootstrap_and_report, daemon=True).start()


def start_background_refresh():
    interval = max(AUTO_REFRESH_HOURS, 0.25) * 3600

    def _tick():
        refresh()
        t = threading.Timer(interval, _tick)
        t.daemon = True
        t.start()

    t = threading.Timer(interval, _tick)
    t.daemon = True
    t.start()


def get_dashboard_data() -> dict:
    return _CACHE["data"] or _load_bundled_fallback()


def sync_status() -> dict:
    return {"last_sync": _CACHE["last_sync"], "source": _CACHE["source"], "error": _CACHE["error"]}


# ── HTML rendering (template + live data injection) ─────────────────────

def render_dashboard_html() -> str:
    with open(TEMPLATE_PATH, "r", encoding="utf-8") as f:
        template = f.read()
    data_json = json.dumps(get_dashboard_data())
    html = template.replace("__NEA_DATA_JSON__", data_json)
    return html
def render_forecast_lab_html() -> str:
    """The Forecast Lab page is static HTML/JS — it pulls its parameter
    list and forecast results live from the /api/nea-forecast-params and
    /api/nea-forecast endpoints (see app.py), so no data injection is
    needed here."""
    with open(FORECAST_TEMPLATE_PATH, "r", encoding="utf-8") as f:
        return f.read()
def forecast_result_to_dict(fr: "ForecastResult") -> dict:
    """JSON-serializable shape for a ForecastResult, used by the
    /api/nea-forecast endpoint."""
    return {
        "past_labels": fr.past_labels, "past_values": fr.past_values,
        "base_label": fr.base_label, "base_value": fr.base_value,
        "pred_labels": fr.pred_labels, "pred_values": fr.pred_values,
        "pred_lo": fr.pred_lo, "pred_hi": fr.pred_hi, "meta": fr.meta,
    }


# ══════════════════════════════════════════════════════════════════════
#  FORECASTING ENGINE
#  (Linear / Holt / Moving-Average / ARIMA / SARIMA / Hybrid)
# ══════════════════════════════════════════════════════════════════════

# See the long unit-handling comment in unit_economics() below before
# trusting any Rs./unit figure derived from these series.

def _annual_series_from_cache() -> dict:
    """Build the same 8-parameter forecast menu as before, but sourced
    from whatever get_dashboard_data() currently holds (live-synced or
    fallback) rather than a hardcoded snapshot — so the Forecast Lab
    tracks new fiscal years automatically, same as the main dashboard."""
    d = get_dashboard_data()
    ae, cg, sr, fin, sl = d["annualEnergy"], d["consumers"], d["sales"], d["financial"], d["systemLoss"]
    tx, ss = d["transmission"], d["substation"]

    def yrs(y_list):
        return [int(y) for y in y_list]

    return {
        "totalAvailability": {"label": "Total Energy Availability (MU)", "unit": "MU",
                               "years": yrs(ae["years"]), "values": ae.get("total", [])},
        "nationalPeak": {"label": "National Peak Demand (MW)", "unit": "MW",
                          "years": yrs(ae["years"]), "values": ae.get("national_peak", [])},
        "systemLoss": {"label": "System Loss (%)", "unit": "%",
                        "years": yrs(sl["years"]), "values": sl["system"]},
        "totalConsumers": {"label": "Total Consumers (No.)", "unit": "consumers",
                            "years": yrs(cg["years"]), "values": cg["total"]},
        "totalRevenue": {"label": "Total Gross Revenue (Rs. Million)", "unit": "Rs. Million",
                          "years": yrs(sr["years"]), "values": sr["total"]},
        "profitLoss": {"label": "Profit / Loss (Rs. Million)", "unit": "Rs. Million",
                        "years": yrs(fin["years"]), "values": fin["profit_loss"]},
        "transmissionTotal": {"label": "Transmission Lines (Circuit Km)", "unit": "Ckt. Km",
                               "years": None, "fy_labels": tx["years"], "values": tx["total"]},
        "substationCap": {"label": "Substation Capacity (MVA)", "unit": "MVA",
                           "years": None, "fy_labels": ss["years"], "values": ss["capacity"]},
    }


def _monthly_series_from_cache() -> dict:
    d = get_dashboard_data()
    eb = d["energyBalanceMonthly"]
    fy_order = sorted(eb.keys())  # chronological if FY strings sort correctly
    labels, values = [], []
    for fy in fy_order:
        entry = eb[fy]
        months = entry.get("months", _MONTHS_BS)
        demand = entry.get("system_demand", [])
        for m, v in zip(months, demand):
            labels.append(f"{fy} {m}")
            values.append(v)
    return {"monthlySystemDemand": {
        "label": "Monthly System Energy Demand (GWh) — seasonal, period=12",
        "unit": "GWh", "labels": labels, "values": values, "season_length": 12,
    }}


def _viable_models_for_values(x_hist, values, monthly, season_length=12):
    """Which models are worth offering for THIS parameter's specific
    history. This is a FAST length-based heuristic, not a real fit —
    actually fitting every candidate model (ARIMA/SARIMA in particular
    do a multi-combination grid search) for every parameter on every
    Forecast Lab page load was far too slow and could time out the
    params endpoint entirely, leaving the dropdowns empty. The real
    protection against a model that fails or spikes on this specific
    data is the automatic fallback chain + spike guard already built
    into _run_single_model(), which only runs once, when the person
    actually clicks Run Forecast — this function just keeps obviously
    hopeless choices (e.g. SARIMA on an 8-point annual series) off the
    menu to begin with. Linear Regression is always included — it
    only needs 2 distinct points and can't diverge."""
    n = len(values)
    ok = ["linear"]
    if n >= 3:
        ok.append("moving")
    if n >= 4:
        ok.append("holt")
    if n >= 6:
        ok.append("arima")
        ok.append("hybrid")
    if monthly and n >= season_length * 2:
        ok.append("sarima")
    return ok


def forecast_param_choices():
    """Dropdown-ready list of every forecastable parameter (annual +
    the one seasonal monthly series), each tagged with whether it's
    monthly (so the caller knows which run_forecast(monthly=) flag to
    use) and with `models`: the list of models that actually run
    cleanly for that specific parameter's history, so the Forecast
    Lab's Model dropdown only ever offers a real, loadable option."""
    out = []
    for k, v in _annual_series_from_cache().items():
        values, span = _clean_series(v["values"])
        if values is None:
            continue
        x_hist = v["years"][span[0]:span[1] + 1] if v.get("years") else list(range(len(values)))
        out.append({"label": v["label"], "value": k, "monthly": False,
                     "models": _viable_models_for_values(x_hist, values, False)})
    for k, v in _monthly_series_from_cache().items():
        values, span = _clean_series(v["values"])
        if values is None:
            continue
        season_length = v.get("season_length", 12)
        out.append({"label": v["label"], "value": k, "monthly": True,
                     "models": _viable_models_for_values(list(range(len(values))), values, True, season_length)})
    return out


MODEL_CHOICES_ANNUAL = [
    {"label": "Linear Regression", "value": "linear"},
    {"label": "Holt Exponential Smoothing (trend)", "value": "holt"},
    {"label": "Moving Average (3-yr)", "value": "moving"},
    {"label": "ARIMA (auto order, AIC-selected)", "value": "arima"},
    {"label": "Hybrid (Linear + ARIMA ensemble)", "value": "hybrid"},
]
MODEL_CHOICES_MONTHLY = MODEL_CHOICES_ANNUAL + [
    {"label": "SARIMA (seasonal, period=12)", "value": "sarima"},
]


def _fit_metrics(actual, fitted) -> dict:
    """R² / RMSE / MAE / MAPE between the historical series and each
    model's in-sample fitted values. `fitted` may contain None/NaN for
    warm-up periods (e.g. ARIMA's differencing order, a moving-average
    window) — those points are dropped, matching the original demo's
    calcMetrics() behaviour, not treated as zero-error."""
    pairs = []
    for a, f in zip(actual, fitted):
        if f is None or a is None:
            continue
        try:
            fv = float(f)
        except (TypeError, ValueError):
            continue
        if np.isnan(fv):
            continue
        pairs.append((float(a), fv))
    if len(pairs) < 2:
        return {"r2": None, "rmse": None, "mae": None, "mape": None}
    a = np.array([p[0] for p in pairs])
    f = np.array([p[1] for p in pairs])
    resid = a - f
    ss_res = float(np.sum(resid ** 2))
    ss_tot = float(np.sum((a - a.mean()) ** 2))
    r2 = None if ss_tot == 0 else round(1 - ss_res / ss_tot, 4)
    rmse = round(float(np.sqrt(ss_res / len(a))), 4)
    mae = round(float(np.mean(np.abs(resid))), 4)
    nz = a != 0
    mape = round(float(np.mean(np.abs(resid[nz] / a[nz]))) * 100, 2) if nz.any() else None
    return {"r2": r2, "rmse": rmse, "mae": mae, "mape": mape}


def _clean_series(raw_vals):
    """Trim a component series to its first..last non-None span and
    linearly interpolate any internal gaps, so a category column with a
    blank cell or two (common in a hand-maintained sheet) still gets a
    clean numeric array to forecast on. Returns (values, (first_idx,
    last_idx)) against the ORIGINAL indices, or (None, None) if there's
    not enough real data (fewer than 3 usable points) to fit anything."""
    n = len(raw_vals)
    first = next((i for i in range(n) if raw_vals[i] is not None), None)
    last = next((i for i in range(n - 1, -1, -1) if raw_vals[i] is not None), None)
    if first is None or last is None or (last - first) < 2:
        return None, None
    vals = [float(v) if v is not None else None for v in raw_vals[first:last + 1]]
    for i, v in enumerate(vals):
        if v is None:
            prev_i = max(j for j in range(i - 1, -1, -1) if vals[j] is not None)
            next_i = min(j for j in range(i + 1, len(vals)) if vals[j] is not None)
            frac = (i - prev_i) / (next_i - prev_i)
            vals[i] = vals[prev_i] + frac * (vals[next_i] - vals[prev_i])
    return vals, (first, last)


def _linear_forecast(x_hist, y_hist, n_ahead):
    x = np.asarray(x_hist, dtype=float)
    y = np.asarray(y_hist, dtype=float)
    slope, intercept = np.polyfit(x, y, 1)
    last_x = x[-1]
    fitted = [float(slope * xi + intercept) for xi in x]
    pred = [float(slope * (last_x + i) + intercept) for i in range(1, n_ahead + 1)]
    return pred, \
        {"model": "Linear Regression", "slope": round(float(slope), 4), "intercept": round(float(intercept), 2)}, \
        fitted


def _moving_avg_forecast(y_hist, n_ahead, window=3):
    y = list(y_hist)
    window = min(window, len(y))
    fitted = [float(sum(y[max(0, i - window + 1):i + 1]) / len(y[max(0, i - window + 1):i + 1]))
              for i in range(len(y))]
    ma = sum(y[-window:]) / window
    trend = 0.0
    if len(y) >= window + 1:
        prev_ma = sum(y[-window - 1:-1]) / window
        trend = ma - prev_ma
    pred = [float(ma + trend * i) for i in range(1, n_ahead + 1)]
    return pred, \
        {"model": f"Moving Average ({window}-period)", "last_ma": round(ma, 2), "trend_per_step": round(trend, 2)}, \
        fitted


def _holt_forecast(y_hist, n_ahead):
    y = pd.Series(y_hist, dtype=float)
    fit = Holt(y, initialization_method="estimated").fit(optimized=True)
    fc = fit.forecast(n_ahead)
    fitted = [float(v) for v in fit.fittedvalues.values]
    return [float(v) for v in fc.values], \
        {"model": "Holt Exponential Smoothing",
         "alpha": round(float(fit.params["smoothing_level"]), 3),
         "beta": round(float(fit.params["smoothing_trend"]), 3)}, \
        fitted


def _best_arima(y_hist, max_p=2, max_d=1, max_q=2):
    y = pd.Series(y_hist, dtype=float)
    best_aic, best_order, best_fit = np.inf, (1, 1, 0), None
    for p, d, q in itertools.product(range(max_p + 1), range(max_d + 1), range(max_q + 1)):
        if p == 0 and q == 0:
            continue
        try:
            fit = ARIMA(y, order=(p, d, q)).fit()
            if np.isfinite(fit.aic) and fit.aic < best_aic:
                best_aic, best_order, best_fit = fit.aic, (p, d, q), fit
        except Exception:
            continue
    if best_fit is None:
        best_fit = ARIMA(y, order=(1, 1, 0)).fit()
        best_order, best_aic = (1, 1, 0), best_fit.aic
    return best_fit, best_order, best_aic


def _arima_forecast(y_hist, n_ahead):
    fit, order, aic = _best_arima(y_hist)
    fc = fit.get_forecast(n_ahead)
    mean = [float(v) for v in fc.predicted_mean.values]
    ci = fc.conf_int(alpha=0.20)
    lo = [float(v) for v in ci.iloc[:, 0].values]
    hi = [float(v) for v in ci.iloc[:, 1].values]
    fitted = [float(v) if not np.isnan(v) else None for v in fit.fittedvalues.values]
    return mean, {"model": f"ARIMA{order}", "aic": round(float(aic), 2)}, lo, hi, fitted


def _sarima_forecast(y_hist, n_ahead, season_length=12):
    y = pd.Series(y_hist, dtype=float)
    best_aic, best_spec, best_fit = np.inf, None, None
    for order, sorder in itertools.product(
        [(1, 1, 0), (0, 1, 1), (1, 1, 1)],
        [(1, 0, 0, season_length), (0, 1, 1, season_length), (1, 1, 0, season_length)],
    ):
        try:
            fit = SARIMAX(y, order=order, seasonal_order=sorder,
                           enforce_stationarity=False, enforce_invertibility=False).fit(disp=False)
            if np.isfinite(fit.aic) and fit.aic < best_aic:
                best_aic, best_spec, best_fit = fit.aic, (order, sorder), fit
        except Exception:
            continue
    if best_fit is None:
        best_fit = SARIMAX(y, order=(1, 1, 0), seasonal_order=(1, 0, 0, season_length),
                            enforce_stationarity=False, enforce_invertibility=False).fit(disp=False)
        best_spec = ((1, 1, 0), (1, 0, 0, season_length))
        best_aic = best_fit.aic
    fc = best_fit.get_forecast(n_ahead)
    mean = [float(v) for v in fc.predicted_mean.values]
    ci = fc.conf_int(alpha=0.20)
    lo = [float(v) for v in ci.iloc[:, 0].values]
    hi = [float(v) for v in ci.iloc[:, 1].values]
    order, sorder = best_spec
    fitted = [float(v) if not np.isnan(v) else None for v in best_fit.fittedvalues.values]
    return mean, {"model": f"SARIMA{order}x{sorder}", "aic": round(float(best_aic), 2)}, lo, hi, fitted


def _hybrid_forecast(x_hist, y_hist, n_ahead):
    lin_vals, lin_meta, lin_fitted = _linear_forecast(x_hist, y_hist, n_ahead)
    ar_vals, ar_meta, lo, hi, ar_fitted = _arima_forecast(y_hist, n_ahead)
    blended = [float((a + b) / 2) for a, b in zip(lin_vals, ar_vals)]
    fitted = []
    for a, b in zip(lin_fitted, ar_fitted):
        if a is not None and b is not None:
            fitted.append(float((a + b) / 2))
        elif a is not None:
            fitted.append(float(a))
        else:
            fitted.append(b)
    meta = {"model": f"Hybrid (Linear + {ar_meta['model']}, averaged)",
            "linear_slope": lin_meta["slope"], "arima_aic": ar_meta["aic"]}
    return blended, meta, fitted


@dataclass
class ForecastResult:
    past_labels: list
    past_values: list
    base_label: str
    base_value: float
    pred_labels: list
    pred_values: list
    pred_lo: list = field(default_factory=list)
    pred_hi: list = field(default_factory=list)
    meta: dict = field(default_factory=dict)


def _is_reasonable_forecast(values, pred):
    """Reject a forecast that jumps far more violently than the
    historical series ever did — the usual symptom of an ill-fitting
    ARIMA/SARIMA order or an unstable Holt trend on a short/noisy
    series, not a genuine signal. Compares each forecasted step-to-step
    change against the largest and typical step-to-step change actually
    observed in history, with generous headroom for a real accelerating
    trend. Used to trigger the model fallback chain in
    _run_single_model(), not to reject a forecast outright."""
    if not pred:
        return False
    try:
        pred_f = [float(v) for v in pred]
    except (TypeError, ValueError):
        return False
    if any(not np.isfinite(v) for v in pred_f):
        return False
    vals = [float(v) for v in values if v is not None]
    if len(vals) < 2:
        return True
    hist_diffs = [abs(vals[i] - vals[i - 1]) for i in range(1, len(vals))]
    max_hist_step = max(hist_diffs) if hist_diffs else 0.0
    typical_step = float(np.median(hist_diffs)) if hist_diffs else 0.0
    last_val = abs(vals[-1])
    cap = max(max_hist_step * 4, typical_step * 8, last_val * 2, 1e-6)
    chain = [vals[-1]] + pred_f
    step_changes = [abs(chain[i] - chain[i - 1]) for i in range(1, len(chain))]
    return all(sv <= cap for sv in step_changes)


# Fallback order tried, in turn, whenever the requested model either
# raises an exception or clears the spike check above. "linear" is the
# final rung — a simple least-squares fit on ≥3 points can't diverge or
# throw, so every parameter is always forecastable by *something*.
_MODEL_FALLBACK_CHAIN = ["holt", "moving", "linear"]


def _fit_one_model(model: str, x_hist, values, n_ahead: int, monthly: bool, season_length: int = 12):
    """Run exactly the requested model, no fallback. Returns
    (pred, meta, lo, hi, fitted); lo/hi are [] for models that don't
    produce a confidence interval. Raises on failure — callers that
    want automatic fallback should go through _run_single_model()."""
    lo, hi = [], []
    if model == "linear":
        pred, meta, fitted = _linear_forecast(x_hist, values, n_ahead)
    elif model == "moving":
        pred, meta, fitted = _moving_avg_forecast(values, n_ahead)
    elif model == "holt":
        pred, meta, fitted = _holt_forecast(values, n_ahead)
    elif model == "arima":
        pred, meta, lo, hi, fitted = _arima_forecast(values, n_ahead)
    elif model == "sarima":
        if not monthly:
            raise ValueError("SARIMA is only offered for the monthly (seasonal) series — "
                              "the annual series are too short and non-seasonal for it to mean anything.")
        pred, meta, lo, hi, fitted = _sarima_forecast(values, n_ahead, season_length)
    elif model == "hybrid":
        pred, meta, fitted = _hybrid_forecast(x_hist, values, n_ahead)
    else:
        raise ValueError(f"Unknown model {model!r}")
    return pred, meta, lo, hi, fitted


def _run_single_model(model: str, x_hist, values, n_ahead: int, monthly: bool, season_length: int = 12):
    """Shared dispatcher used by both run_forecast() (single series) and
    run_composite_forecast() (per-component series inside a stacked
    group), so every parameter — whether it's charted on its own or as
    one slice of a stack — goes through the exact same statsmodels
    fitting code, with the same automatic fallback and anti-spike
    correction. Returns (pred, meta, lo, hi, fitted); lo/hi are []
    for models that don't produce a confidence interval.

    FALLBACK CHAIN: the requested model is tried first. If it raises
    (too little data, a non-convergent fit, an unsupported
    combination like SARIMA on an annual series) or its forecast
    clears the spike check in _is_reasonable_forecast(), the next
    model in _MODEL_FALLBACK_CHAIN is tried instead, ending at Linear
    Regression — which only needs 2 distinct points and can't diverge
    — so every parameter always produces *some* forecast rather than
    an error or a wild spike. meta['requested_model'] /
    meta['fallback_used'] record when this happened."""
    chain = [model] + [m for m in _MODEL_FALLBACK_CHAIN if m != model]
    pred = meta = fitted = None
    lo, hi = [], []
    last_err = None
    for m in chain:
        try:
            pred, meta, lo, hi, fitted = _fit_one_model(m, x_hist, values, n_ahead, monthly, season_length)
        except Exception as e:
            last_err = e
            continue
        if m == "linear" or _is_reasonable_forecast(values, pred):
            if m != model:
                meta = {**meta, "requested_model": model, "fallback_used": m}
            break
    else:
        raise last_err or ValueError(f"Could not fit any forecasting model for this parameter.")

    # ── CONTINUITY CORRECTION ───────────────────────────────────────
    # Holt/ARIMA/SARIMA forecast forward from their own internal
    # smoothed level, not from the raw last actual observation. When
    # that internal level doesn't exactly match the last real data
    # point (e.g. a volatile last year), the forecast trajectory is
    # still internally consistent, but the *chart* — which draws the
    # raw history line up to the real base value and then starts the
    # forecast line — shows an artificial jump right at the forecast
    # start, even though nothing is actually wrong with the model.
    # Anchor the whole forecast to the last real observation by
    # shifting every predicted (and CI) point by the model's own
    # last in-sample residual, so the forecast line always leaves
    # from exactly where the history line ends — the forecast's
    # first point joins the base value with no jump.
    last_actual = values[-1] if len(values) else None
    last_fitted = fitted[-1] if fitted else None
    if last_actual is not None and last_fitted is not None:
        try:
            offset = float(last_actual) - float(last_fitted)
        except (TypeError, ValueError):
            offset = 0.0
        if offset:
            pred = [p + offset for p in pred]
            if lo:
                lo = [v + offset for v in lo]
            if hi:
                hi = [v + offset for v in hi]
    return pred, meta, lo, hi, fitted


def run_forecast(param_key: str, model: str, n_ahead: int, monthly: bool = False) -> ForecastResult:
    if not get_dashboard_data():
        raise ValueError("NEA operational data is not available right now. Please check back later.")
    n_ahead = max(1, min(int(n_ahead), 20))
    if monthly:
        series = _monthly_series_from_cache()[param_key]
        raw_labels = series["labels"]
        raw_values = series["values"]
        season_length = series.get("season_length", 12)
    else:
        series = _annual_series_from_cache()[param_key]
        raw_values = series["values"]
        if series.get("years"):
            raw_labels = [str(y) for y in series["years"]]
        else:
            raw_labels = series["fy_labels"]
        season_length = 12

    # Trim to the real (first..last non-None) span and interpolate any
    # internal gaps — a hand-maintained sheet with a blank cell or two
    # shouldn't crash every model, including the guaranteed Linear
    # fallback, which needs plain floats.
    values, span = _clean_series(raw_values)
    if values is None:
        raise ValueError("Not enough data to forecast this parameter yet (fewer than 3 usable points).")
    first, last = span
    labels = raw_labels[first:last + 1]
    x_hist = series["years"][first:last + 1] if (not monthly and series.get("years")) else list(range(len(values)))

    if monthly:
        pred_labels = [f"+{i} mo" for i in range(1, n_ahead + 1)]
    elif series.get("years"):
        last_year = x_hist[-1]
        pred_labels = [str(last_year + i) for i in range(1, n_ahead + 1)]
    else:
        pred_labels = [f"FY+{i}" for i in range(1, n_ahead + 1)]

    pred, meta, lo, hi, fitted = _run_single_model(model, x_hist, values, n_ahead, monthly, season_length)
    meta = {**meta, **_fit_metrics(values, fitted)}

    return ForecastResult(
        past_labels=labels, past_values=[float(v) for v in values],
        base_label=labels[-1], base_value=float(values[-1]),
        pred_labels=pred_labels, pred_values=pred,
        pred_lo=lo, pred_hi=hi, meta=meta,
    )


# ══════════════════════════════════════════════════════════════════════
#  COMPOSITE (STACKED / MULTI-COMPONENT) FORECASTING
#
#  For a parameter that's really made of several categories that sum
#  to a total — Consumers by Category, Revenue by Consumer Category,
#  the annual Generation Mix (NEA Own / NEA Subsidiary / IPP / India
#  Import), and the monthly Energy Mix — the Forecast Lab shouldn't
#  just forecast the aggregate as one line. It forecasts EACH component
#  series independently (own model fit, own confidence band, own
#  accuracy stats) and then stacks the forecasted components, so the
#  final "scenario" is a real stacked chart, not a single blended line.
# ══════════════════════════════════════════════════════════════════════

def _annual_composite_defs() -> dict:
    """Stack-able annual parameter groups, sourced from the same live
    cache as _annual_series_from_cache(). Component keys/labels for the
    consumer and revenue breakdowns are discovered dynamically from
    whatever category columns are actually in the sheet, so a renamed
    or newly added category (e.g. a new "Irrigation" consumer class)
    shows up automatically — nothing here is hardcoded to today's
    column headers except the four fixed Generation Mix sources."""
    d = get_dashboard_data()
    ae, cg, sr = d.get("annualEnergy", {}), d.get("consumers", {}), d.get("sales", {})
    out = {}

    gen_components = []
    for key, label in (("nea_own", "NEA Own"), ("nea_sub", "NEA Subsidiary"),
                        ("ipp", "IPP"), ("india", "India Import")):
        vals = ae.get(key)
        if vals and any(v is not None for v in vals):
            gen_components.append({"key": key, "label": label, "values": vals})
    if gen_components and ae.get("years"):
        out["generationMixAnnual"] = {
            "label": "Generation Mix — Annual", "unit": "MU", "monthly": False,
            "years": [int(y) for y in ae["years"]], "components": gen_components,
            "reported_total": ae.get("total", []),
        }

    if cg.get("years") and cg.get("categories"):
        cons_components = [{"key": _slug(k) or f"cat{i}", "label": k, "values": v}
                            for i, (k, v) in enumerate(cg["categories"].items())
                            if v and any(x is not None for x in v)]
        if cons_components:
            out["consumersByCategory"] = {
                "label": "Consumers by Category", "unit": "consumers", "monthly": False,
                "years": [int(y) for y in cg["years"]], "components": cons_components,
                "reported_total": cg.get("total", []),
            }

    if sr.get("years") and sr.get("categories"):
        rev_components = [{"key": _slug(k) or f"cat{i}", "label": k, "values": v}
                           for i, (k, v) in enumerate(sr["categories"].items())
                           if v and any(x is not None for x in v)]
        if rev_components:
            out["revenueByCategory"] = {
                "label": "Revenue by Consumer Category", "unit": "Rs. Million", "monthly": False,
                "years": [int(y) for y in sr["years"]], "components": rev_components,
                "reported_total": sr.get("total", []),
            }

    return out


def _monthly_composite_defs() -> dict:
    """The one stack-able monthly group: the generation-source mix
    (IPP / NEA Subsidiary / NEA ROR-PROR / NEA Storage / NEA Solar /
    Thermal / Import) behind the monthly Energy Balance sheet, flattened
    across every fiscal year the same way _monthly_series_from_cache()
    flattens the single system-demand series, so it lines up with the
    same SARIMA(season=12) treatment."""
    d = get_dashboard_data()
    eb = d.get("energyBalanceMonthly", {})
    fy_order = sorted(eb.keys())
    comp_defs = [("ipp", "IPP"), ("nea_sub", "NEA Subsidiary"), ("nea_ror", "NEA ROR/PROR"),
                 ("nea_storage", "NEA Storage"), ("nea_solar", "NEA Solar"),
                 ("thermal", "Thermal"), ("import", "Import")]
    labels, comp_vals, reported_total = [], {k: [] for k, _ in comp_defs}, []
    for fy in fy_order:
        entry = eb[fy]
        months = entry.get("months", _MONTHS_BS)
        demand = entry.get("system_demand", [])
        for i in range(len(months)):
            labels.append(f"{fy} {months[i]}")
            for k, _ in comp_defs:
                series = entry.get(k, [])
                comp_vals[k].append(series[i] if i < len(series) else None)
            reported_total.append(demand[i] if i < len(demand) else None)

    components = [{"key": k, "label": lbl, "values": comp_vals[k]}
                  for k, lbl in comp_defs if any(v is not None for v in comp_vals[k])]
    out = {}
    if components and labels:
        out["energyMixMonthly"] = {
            "label": "Monthly Energy Mix (Generation Sources)", "unit": "GWh", "monthly": True,
            "labels": labels, "components": components, "reported_total": reported_total,
            "season_length": 12,
        }
    return out


def composite_param_choices():
    """Dropdown-ready list of every stack-able (composite) parameter,
    parallel to forecast_param_choices() for single series. `models`
    is the intersection of what's viable for every component in the
    group — the composite forecast fits the same requested model to
    each component, so a model only belongs on this dropdown if it
    will actually run for all of them, not just some."""
    _ORDER = ["linear", "holt", "moving", "arima", "hybrid", "sarima"]

    def _group_models(v):
        common = None
        for comp in v["components"]:
            values, span = _clean_series(comp["values"])
            if values is None:
                continue
            x_hist = list(range(len(values)))
            viable = set(_viable_models_for_values(x_hist, values, v["monthly"], v.get("season_length", 12)))
            common = viable if common is None else (common & viable)
        return [m for m in _ORDER if m in common] if common else ["linear"]

    annual = [{"label": v["label"], "value": k, "monthly": False, "unit": v["unit"],
               "models": _group_models(v)}
              for k, v in _annual_composite_defs().items()]
    monthly = [{"label": v["label"], "value": k, "monthly": True, "unit": v["unit"],
                "models": _group_models(v)}
               for k, v in _monthly_composite_defs().items()]
    return annual + monthly


def run_composite_forecast(composite_key: str, model: str, n_ahead: int) -> dict:
    """Forecast every component of a stacked group independently, then
    sum the forecasted components period-by-period to build the final
    stacked scenario. Returns a JSON-friendly dict (not a dataclass,
    since the shape — a list of components plus one aggregate — doesn't
    fit ForecastResult)."""
    if not get_dashboard_data():
        raise ValueError("NEA operational data is not available right now. Please check back later.")
    n_ahead = max(1, min(int(n_ahead), 20))

    annual_defs = _annual_composite_defs()
    monthly_defs = _monthly_composite_defs()
    if composite_key in annual_defs:
        cdef = annual_defs[composite_key]
        monthly = False
        years = cdef["years"]
        past_labels = [str(y) for y in years]
        pred_labels = [str(years[-1] + i) for i in range(1, n_ahead + 1)]
        season_length = 12
    elif composite_key in monthly_defs:
        cdef = monthly_defs[composite_key]
        monthly = True
        past_labels = cdef["labels"]
        pred_labels = [f"+{i} mo" for i in range(1, n_ahead + 1)]
        season_length = cdef.get("season_length", 12)
    else:
        raise ValueError(f"Unknown composite parameter {composite_key!r}")

    if model == "sarima" and not monthly:
        raise ValueError("SARIMA is only offered for monthly composite parameters.")

    components_out = []
    agg_pred = [0.0] * n_ahead
    agg_lo = [0.0] * n_ahead
    agg_hi = [0.0] * n_ahead
    have_ci = False

    for comp in cdef["components"]:
        raw_vals = comp["values"]
        vals, span = _clean_series(raw_vals)
        if vals is None or len(vals) < 3:
            continue  # not enough real data points to fit this component
        first, _last = span
        this_x = ([years[first + i] for i in range(len(vals))] if not monthly
                   else list(range(len(vals))))
        try:
            pred, meta, lo, hi, fitted = _run_single_model(model, this_x, vals, n_ahead, monthly, season_length)
        except Exception as exc:
            # one bad component (e.g. too short for the chosen model)
            # shouldn't blank out the whole stack — skip it and keep going
            components_out.append({
                "key": comp["key"], "label": comp["label"],
                "past_values": [float(v) if v is not None else None for v in raw_vals],
                "pred_values": [], "pred_lo": [], "pred_hi": [],
                "meta": {"model": model, "error": str(exc)},
            })
            continue
        meta = {**meta, **_fit_metrics(vals, fitted)}
        if lo and hi:
            have_ci = True
        components_out.append({
            "key": comp["key"], "label": comp["label"],
            "past_values": [float(v) if v is not None else None for v in raw_vals],
            "pred_values": [float(v) for v in pred],
            "pred_lo": [float(v) for v in lo] if lo else [],
            "pred_hi": [float(v) for v in hi] if hi else [],
            "meta": meta,
        })
        for i in range(n_ahead):
            agg_pred[i] += pred[i]
            if lo and hi:
                agg_lo[i] += lo[i]
                agg_hi[i] += hi[i]

    reported_total = cdef.get("reported_total") or []
    return {
        "key": composite_key, "label": cdef["label"], "unit": cdef["unit"], "monthly": monthly,
        "past_labels": past_labels, "pred_labels": pred_labels,
        "components": components_out,
        "aggregate": {
            "label": f"{cdef['label']} — Total (sum of forecasted components)",
            "pred_values": agg_pred,
            "pred_lo": agg_lo if have_ci else [],
            "pred_hi": agg_hi if have_ci else [],
            "reported_past_values": [float(v) if v is not None else None for v in reported_total],
        },
    }


def unit_economics():
    """Rs./unit (Rs./kWh) rates, computed live from get_dashboard_data().

    UNIT HANDLING: 1 "Unit" of electricity = 1 kWh, so 1 Million Units
    (MU) = 1 GWh — MU and GWh are the same physical quantity under
    different names, no conversion factor between them. For a per-unit
    price in Rs./kWh: Rs./unit = (Rs. Million) / (MU), because both
    numerator and denominator carry the same implicit ×10^6 (Rs. Million
    → Rs. is ×10^6, MU → units is ×10^6), which cancels exactly. No
    further scaling is applied — years with 0 reported MU (no
    import/export that year) are skipped for that rate rather than
    shown as a bogus 0.00 or a divide-by-zero.
    """
    d = get_dashboard_data()
    fin, ae = d["financial"], d["annualEnergy"]
    avail_by_year = dict(zip(ae["years"], ae.get("total", [])))
    revenue_by_year = dict(zip(fin["years"], fin["revenue"]))

    out = {"fy": [], "import_rate_rs_per_unit": [], "export_rate_rs_per_unit": [],
           "avg_revenue_rate_rs_per_unit": []}
    for i, y in enumerate(fin["years"]):
        out["fy"].append(y)
        imp_mu = fin["import_mu"][i] if i < len(fin["import_mu"]) else None
        exp_mu = fin["export_mu"][i] if i < len(fin["export_mu"]) else None
        imp_rs = fin["import_rs"][i] if i < len(fin["import_rs"]) else None
        exp_rs = fin["export_rs"][i] if i < len(fin["export_rs"]) else None
        out["import_rate_rs_per_unit"].append(round(imp_rs / imp_mu, 2) if imp_mu else None)
        out["export_rate_rs_per_unit"].append(round(exp_rs / exp_mu, 2) if exp_mu else None)
        avail = avail_by_year.get(y)
        revenue = revenue_by_year.get(y)
        out["avg_revenue_rate_rs_per_unit"].append(
            round(revenue / avail, 2) if (avail and revenue) else None)
    return out
# ══════════════════════════════════════════════════════════════════════
# ADMIN PANEL HELPERS  (add these to the bottom of NEA.py,
# just before the "if __name__ == '__main__':" block)
# ══════════════════════════════════════════════════════════════════════

def load_from_path(path: str) -> bool:
    """Load NEA data from a specific file path (for admin workbook upload).
    Copies the file to the cache path so it persists across restarts.
    Returns True on success, False on failure (error is stored in cache)."""
    with _lock:
        try:
            parsed = parse_workbook(path)
            data = build_dashboard_data(parsed)
            import shutil
            shutil.copy(path, CACHE_WORKBOOK_PATH)
            _CACHE.update(
                data=data,
                parsed=parsed,
                last_sync=datetime.now().strftime("%Y-%m-%d %H:%M"),
                source="Admin upload",
                error=None,
            )
            return True
        except Exception as exc:
            traceback.print_exc()
            _CACHE["error"] = str(exc)
            return False


def get_admin_status() -> dict:
    """Return a JSON-friendly status dict for the admin panel.
    Safe to call from any thread — returns a snapshot of current state."""
    with _lock:
        return {
            "last_sync": _CACHE["last_sync"],
            "source": _CACHE["source"],
            "error": _CACHE["error"],
            "has_data": _CACHE["data"] is not None,
        }

# ── Smoke test ────────────────────────────────────────────────────────

if __name__ == "__main__":
    ok = refresh()
    print("live refresh ok:", ok, sync_status())
    d = get_dashboard_data()
    print("years:", d["systemLoss"]["years"])
    print("kpi:", d["kpi"])
    fr = run_forecast("systemLoss", "arima", 5)
    print("ARIMA forecast:", fr.pred_labels, fr.pred_values, fr.meta)
    fr2 = run_forecast("systemLoss", "hybrid", 5)
    print("Hybrid forecast:", fr2.pred_values, fr2.meta)
    print("unit economics:", unit_economics())
